#!/usr/bin/env python3
"""
This script reads a CSV file containing file paths, opens each file,
and extracts metadata values for ms.author, description, and pivot from the front matter.
It then adds these metadata columns to the CSV and exports the enhanced data.

The script creates:
1. Main CSV file with comma-separated pivot_groups column

Optional: Merges MERGE_COLUMNS from existing Excel file if MERGE_EXISTING and EXISTING_EXCEL_FILE is set.
Optional: Merges engagement metrics if MERGE_ENGAGEMENT and ENGAGEMENT_FILE is set.

Usage:
    python add-metadata.py
"""

import pandas as pd
import os
import dotenv
import sys
import logging
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Any
from utils.file_utils import extract_front_matter, resolve_file_path, load_pivot_mapping, resolve_pivot_groups, parse_metadata_from_content
from utils.url_normalizer import normalize_url
from utils.config_utils import setup_logging, load_configuration
from utils.stats_utils import generate_statistics
from utils.excel_utils import merge_external_data
try:
    import openpyxl
except ImportError:
    openpyxl = None

# Configure logging
logger = logging.getLogger(__name__)

# Load environment variables from .env file
dotenv.load_dotenv()



def process_metadata_extraction(df: pd.DataFrame, config: Dict[str, Any], pivot_mapping: Dict[str, Any]) -> Tuple[int, int]:
    """
    Process each row in the dataframe to extract metadata from markdown files.
    
    Args:
        df: The dataframe to process
        config: Configuration dictionary containing paths and settings
        pivot_mapping: Pivot group mappings for resolving zone pivot groups
        
    Returns:
        Tuple of (processed_files, found_files) counts
    """
    DEBUG = config.get('DEBUG', config.get('debug', False))
    base_path = config['base_path']
    metadata_fields = config['metadata_fields']
    has_pivot_field = config['has_pivot_field']
    metadata_flags = config['metadata_flags']
    
    total_rows = len(df)
    processed_files = 0
    found_files = 0
    
    for index, row in df.iterrows():
        href = row.get('filename', '')
        
        if DEBUG and index % 50 == 0:  # Progress indicator
            logger.debug(f"Processing row {index + 1}/{total_rows}")
        
        # Resolve the file path
        file_path = resolve_file_path(href, base_path)
        
        if file_path:
            df.at[index, 'file_found'] = True
            found_files += 1
            
            # Extract metadata from the file
            metadata = extract_front_matter(file_path)
            
            # Extract configured metadata fields
            for field in metadata_fields:
                if field in metadata:
                    df.at[index, field] = metadata[field]
            
            # Handle pivot groups (only if zone_pivot_groups is in metadata fields)
            if has_pivot_field:
                pivot_group_ids = None
                if 'zone_pivot_groups' in metadata:
                    df.at[index, 'pivot_id'] = metadata['zone_pivot_groups']
                    pivot_group_ids = metadata['zone_pivot_groups']
                
                # Set pivot group data (always use comma-separated for main file)
                if pivot_group_ids:
                    pivot_groups = resolve_pivot_groups(pivot_group_ids, pivot_mapping)
                    
                    # Set has_pivots flag
                    df.at[index, 'has_pivots'] = True
                    
                    # Always set the comma-separated column for main file
                    df.at[index, 'pivot_groups'] = ','.join(pivot_groups) if pivot_groups else ""
                
            # Handle metadata fields with flag logic
            for field_name, flag_name in metadata_flags.items():
                if field_name in metadata:
                    flag_found = False
                    custom_data = metadata[field_name]
                    
                    if isinstance(custom_data, str):
                        # If field is a string, check if it contains the flag
                        flag_found = flag_name in custom_data.lower()
                    elif isinstance(custom_data, list):
                        # If field is a list, check if any item contains the flag
                        flag_found = any(flag_name in str(item).lower() for item in custom_data)
                    elif isinstance(custom_data, dict):
                        # If field is a dict, check if any value contains the flag
                        flag_found = any(flag_name in str(value).lower() for value in custom_data.values())
                    
                    # Set the flag
                    df.at[index, flag_name] = flag_found
            
            processed_files += 1
    
    return processed_files, found_files





def add_metadata_to_csv() -> None:
    """
    Main function to read CSV, extract metadata, and create enhanced CSV.
    
    This orchestrates the complete workflow:
    1. Load configuration from environment variables
    2. Setup logging 
    3. Load pivot mappings
    4. Read input CSV
    5. Process metadata extraction
    6. Merge external data if configured
    7. Generate statistics and save output
    
    Raises:
        SystemExit: If configuration is invalid or processing fails
    """
    try:
        # Load and validate configuration
        config = load_configuration()
        DEBUG = config.get('DEBUG', config.get('debug', False))
        
        # Setup logging
        setup_logging(DEBUG)
        
        # Load pivot mapping
        pivot_mapping = load_pivot_mapping(config['pivot_map_file'])
        if pivot_mapping:
            if DEBUG:
                logger.debug(f"Loaded {len(pivot_mapping)} pivot groups from: {config['pivot_map_file']}")
        else:
            if DEBUG:
                logger.debug("No pivot mapping loaded")
        
        # Read the CSV file
        df = pd.read_csv(config['input_path'])
        
        # Add new columns for metadata dynamically
        for field in config['metadata_fields']:
            df[field] = ""
        
        # Add pivot-related columns only if zone_pivot_groups is in metadata fields
        if config['has_pivot_field']:
            df['pivot_id'] = ""
            df['has_pivots'] = False
            df['pivot_groups'] = ""
        
        # Add metadata flag columns
        for flag_name in config['metadata_flags'].values():
            df[flag_name] = False
        
        # Add system columns
        df['file_found'] = False
        
        # Process metadata extraction for all rows
        processed_files, found_files = process_metadata_extraction(df, config, pivot_mapping)
        
        # Merge existing Excel file data using extracted function
        df = merge_external_data(df, config)


        # Save the main enhanced CSV (always with comma-separated pivot_groups)
        df.to_csv(config['output_path'], index=False)
        
        # Generate and display statistics
        generate_statistics(df, config, processed_files, found_files)
        
        return df  # Return main dataframe only
        
    except (ValueError, FileNotFoundError) as e:
        logger.error(f"Configuration error: {e}")
        return None
    except Exception as e:
        logger.error(f"Unexpected error during metadata processing: {e}")
        return None

def create_excel_analysis(csv_file_path=None, output_file_name=None):
    """
    Create an Excel file with multiple tabs for analysis.
    
    Args:
        csv_file_path (str): Path to the CSV file with content analysis data.
                           If None, looks for toc_with_content.csv in script directory.
        output_file_name (str): Base name for output file. Excel extension will be added.
                              If None, uses "toc_analysis.xlsx"
    """
    # Check if debug mode is enabled
    DEBUG = os.getenv("DEBUG", "False").lower() in ('true', '1', 'yes')
    try:
        import openpyxl
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("Error: openpyxl is required for Excel export. Install with: pip install openpyxl")
        return
    
    # Get the directory of the current script
    script_dir = os.path.dirname(os.path.abspath(__file__))
    
    # Determine input CSV file
    if csv_file_path is None:
        csv_file_path = os.path.join(script_dir, "toc_with_content.csv")
    
    if not os.path.exists(csv_file_path):
        print(f"Error: CSV file {csv_file_path} not found.")
        print("Make sure to run the complete analysis pipeline first.")
        return
    
    # Read the CSV file
    print(f"Reading CSV file: {csv_file_path}")
    df = pd.read_csv(csv_file_path)

    # URL normalization already imported at top of file

    # Merge engagement metrics if available and enabled
    merge_engagement = os.getenv("MERGE_ENGAGEMENT", "False").lower() in ('true', '1', 'yes')
    engagement_file = os.getenv("ENGAGEMENT_FILE")
    if DEBUG:
        print(f"Debug: MERGE_ENGAGEMENT env var = {os.getenv('MERGE_ENGAGEMENT')}")
        print(f"Debug: merge_engagement = {merge_engagement}")
        print(f"Debug: ENGAGEMENT_FILE = {engagement_file}")
        if engagement_file:
            print(f"Debug: File exists? {os.path.exists(engagement_file)}")
    df_engage = None
    if merge_engagement and engagement_file and os.path.exists(engagement_file):
        try:
            print(f"Merging engagement metrics from: {engagement_file}")
            df_engage = pd.read_csv(engagement_file)
            # Only keep relevant columns
            engage_cols = ["Url", "PageViews", "PVMoM", "Visitors", "Engagement"]
            df_engage = df_engage[engage_cols]
            
            # Convert numeric columns from string format (remove commas, percentages and convert to numeric)
            numeric_cols = ["PageViews", "PVMoM", "Visitors"]
            for col in numeric_cols:
                if col in df_engage.columns:
                    # Remove commas and percentage signs, then convert to numeric
                    df_engage[col] = df_engage[col].astype(str).str.replace(',', '').str.replace('%', '')
                    df_engage[col] = pd.to_numeric(df_engage[col], errors='coerce')
                    if DEBUG:
                        print(f"Converted {col} to numeric format (removed commas and %)")
            
            if DEBUG:
                print("Normalizing engagement URLs...")
            df_engage["url_match"] = df_engage["Url"].apply(normalize_url)
            if DEBUG:
                print("\nFirst 5 engagement URLs and their normalized versions:")
                print(df_engage[["Url", "url_match"]].head())
        except Exception as e:
            print(f"Warning: Failed to read engagement file: {e}")
            df_engage = None
            
        if df_engage is not None:
            if "URL" in df.columns:
                if DEBUG:
                    print("\nNormalizing main DataFrame URLs...")
                df["url_match"] = df["URL"].apply(normalize_url)
                if DEBUG:
                    print("\nFirst 5 TOC URLs and their normalized versions:")
                    print(df[["URL", "url_match"]].head())
                    print("\nAttempting merge on 'url_match' column...")
                before_merge = len(df)
                
                # Debug: Show some stats before merge
                if DEBUG:
                    print(f"\nUnique normalized URLs in engagement data: {df_engage['url_match'].nunique()}")
                    print(f"Unique normalized URLs in TOC data: {df['url_match'].nunique()}")
                    print("\nSample of URLs that should match:")
                    sample = df.merge(df_engage, left_on='url_match', right_on='url_match').head()
                    if not sample.empty:
                        print(sample[['URL', 'Url', 'url_match']].head())
                    else:
                        print("No matching URLs found!")
                
                # Perform the merge
                df = df.merge(df_engage.drop(columns=["Url"]), how="left", on="url_match")
                after_merge = len(df)
                
                engagement_matches = df['PageViews'].notna().sum()
                print(f"Engagement data merged: {engagement_matches} URLs matched")
                if DEBUG:
                    print(f"Merge complete. Rows before: {before_merge}, after: {after_merge}")

                # Drop the temporary url_match column as it's no longer needed
                df = df.drop(columns=["url_match"])
                
                # Debug: Show results
                if DEBUG:
                    print("\nSample of merged data:")
                    print(df[['URL', 'PageViews']].head())
            else:
                print("Main DataFrame missing 'URL' column!")
        else:
            if DEBUG:
                print("No engagement data available to merge")
        if DEBUG:
            print(df.head(10))
    
    # Check for pivot columns file
    df_pivots = None
    # pivot_file_path = csv_file_path.replace('.csv', '-pivots.csv')
    # if os.path.exists(pivot_file_path):
    #     print(f"Reading pivot columns file: {pivot_file_path}")
    #     df_pivots = pd.read_csv(pivot_file_path)
    
    
    # Reorder columns to ensure merged columns are in early positions
    merge_columns_config = os.getenv("MERGE_COLUMNS", "URL,Notes,NextGen?,NextGen TOC")
    desired_merge_columns = [col.strip() for col in merge_columns_config.split(',')]
    key_column = desired_merge_columns[0] if desired_merge_columns else 'URL'
    
    special_columns = []
    for col in desired_merge_columns[1:]:  # Skip key column
        if col in df.columns:
            special_columns.append(col)
    
    if special_columns:
        # Define desired column order
        base_columns = ['Parent Path', 'Name', 'filename', key_column]  # First 4 columns (A-D)
        
        # Get all other columns except the base ones and special columns
        other_columns = [col for col in df.columns if col not in base_columns + special_columns]
        
        # Create final column order: base + special + others
        final_column_order = base_columns + special_columns + other_columns
        
        # Only include columns that actually exist in the dataframe
        final_column_order = [col for col in final_column_order if col in df.columns]
        
        # Reorder the dataframe
        df = df[final_column_order]
        if DEBUG:
            print(f"Reordered columns with {', '.join(special_columns)} in early positions")
    
    # Create Excel file path using output file name
    if output_file_name:
        # Remove extension if present and add .xlsx
        base_name = os.path.splitext(output_file_name)[0]
        excel_file_path = os.path.join(script_dir, f"{base_name}.xlsx")
    else:
        excel_file_path = os.path.join(script_dir, "toc_analysis.xlsx")
    
    # Create Excel writer
    with pd.ExcelWriter(excel_file_path, engine='openpyxl') as writer:
        
        # Tab: Complete data
        print("Creating Tab: Complete Data")
        df.to_excel(writer, sheet_name='Complete Data', index=False)
        
        # # Tab: Pivot columns (if available)
        # if df_pivots is not None:
        #     print("Creating Tab: Pivot Columns")
        #     df_pivots.to_excel(writer, sheet_name='Pivot Columns', index=False)
        
        # Tab: Summary statistics
        summary_tab_name = 'Content Summary'
        print(f"Creating Tab: {summary_tab_name}")
        
        # Create summary data
        summary_data = []
        
        # Content type summaries
        if 'has_tabs' in df.columns:
            files_with_tabs = len(df[df['has_tabs'] == True])
            total_tabs = df['tab_count'].sum() if 'tab_count' in df.columns else 0
            summary_data.append(['Content Type', 'Files with Content', 'Total Instances'])
            summary_data.append(['Tabs', files_with_tabs, total_tabs])
        
        if 'has_images' in df.columns:
            files_with_images = len(df[df['has_images'] == True])
            total_images = df['image_count'].sum() if 'image_count' in df.columns else 0
            summary_data.append(['Images', files_with_images, total_images])
        
        if 'has_code_blocks' in df.columns:
            files_with_code_blocks = len(df[df['has_code_blocks'] == True])
            total_code_blocks = df['code_block_count'].sum() if 'code_block_count' in df.columns else 0
            summary_data.append(['Code Blocks', files_with_code_blocks, total_code_blocks])
        
        if 'has_code_refs' in df.columns:
            files_with_code_refs = len(df[df['has_code_refs'] == True])
            total_code_refs = df['code_ref_count'].sum() if 'code_ref_count' in df.columns else 0
            summary_data.append(['Code References', files_with_code_refs, total_code_refs])
        
        if 'portal_steps' in df.columns:
            files_with_portal_steps = len(df[df['portal_steps'] == True])
            summary_data.append(['Portal Steps', files_with_portal_steps, '-'])
        
        # Create summary DataFrame if we have data
        if len(summary_data) > 0:
            summary_df = pd.DataFrame(summary_data[1:], columns=summary_data[0])
            summary_df.to_excel(writer, sheet_name=summary_tab_name, index=False, startrow=0)
        else:
            # Create an empty DataFrame with default columns
            summary_df = pd.DataFrame(columns=['Content Type', 'Files with Content', 'Total Instances'])
            summary_df.to_excel(writer, sheet_name=summary_tab_name, index=False, startrow=0)
        
        # Add metadata summary
        metadata_summary = []
        metadata_summary.append(['Metadata Type', 'Files with Metadata'])
        
        if 'ms.author' in df.columns:
            files_with_author = len(df[df['ms.author'].notna() & (df['ms.author'] != '')])
            metadata_summary.append(['Authors', files_with_author])
        
        if 'ms.topic' in df.columns:
            files_with_topic = len(df[df['ms.topic'].notna() & (df['ms.topic'] != '')])
            metadata_summary.append(['Topics', files_with_topic])
        
        if 'ms.service' in df.columns:
            files_with_service = len(df[df['ms.service'].notna() & (df['ms.service'] != '')])
            # Count unique services
            unique_services = df[df['ms.service'].notna() & (df['ms.service'] != '')]['ms.service'].nunique()
            metadata_summary.append(['Services', f"{files_with_service} files ({unique_services} unique)"])
            
            # Add a breakdown of services
            service_counts = df[df['ms.service'].notna() & (df['ms.service'] != '')]['ms.service'].value_counts()
            for service, count in service_counts.items():
                metadata_summary.append([f"  {service}", count])
        
        if 'description' in df.columns:
            files_with_description = len(df[df['description'] != ''])
            metadata_summary.append(['Descriptions', files_with_description])
        
        if 'has_pivots' in df.columns:
            files_with_pivots = len(df[df['has_pivots'] == True])
            metadata_summary.append(['Pivot Groups', files_with_pivots])
        
        if 'hub-only' in df.columns:
            files_hub_only = len(df[df['hub-only'] == True])
            metadata_summary.append(['Hub-Only', files_hub_only])
        
        # Add metadata summary to the same sheet
        metadata_df = pd.DataFrame(metadata_summary[1:], columns=metadata_summary[0])
        metadata_df.to_excel(writer, sheet_name=summary_tab_name, index=False, startrow=len(summary_df) + 3)
        
        # Add top programming languages if available
        if 'code_languages' in df.columns:
            all_languages = []
            for _, row in df.iterrows():
                if row['code_languages'] and pd.notna(row['code_languages']):
                    languages = [lang.strip() for lang in str(row['code_languages']).split(',')]
                    all_languages.extend(languages)
            
            if all_languages:
                from collections import Counter
                lang_counts = Counter(all_languages)
                
                lang_summary = [['Programming Language', 'File Count']]
                for lang, count in lang_counts.most_common(10):
                    if lang and lang != 'none':
                        lang_summary.append([lang, count])
                
                lang_df = pd.DataFrame(lang_summary[1:], columns=lang_summary[0])
                lang_df.to_excel(writer, sheet_name=summary_tab_name, index=False, 
                               startrow=len(summary_df) + len(metadata_df) + 6)
    
    # Format the Excel file
    wb = openpyxl.load_workbook(excel_file_path)
    
    # Convert data sheets to Excel tables (for filtering and sorting)
    data_sheets = ['Complete Data']
    for sheet_name in data_sheets:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Get the data range
            max_row = ws.max_row
            max_col = ws.max_column
            
            if max_row > 1 and max_col > 0:  # Only create table if there's data
                # Create table reference
                table_ref = f"A1:{ws.cell(max_row, max_col).coordinate}"
                
                # Create table
                from openpyxl.worksheet.table import Table, TableStyleInfo
                table = Table(displayName=f"Table_{sheet_name.replace(' ', '')}", ref=table_ref)
                
                # Add a clean table style with dark blue headers
                style = TableStyleInfo(
                    name="TableStyleMedium16",  # Dark blue headers with clean white rows
                    showFirstColumn=False,
                    showLastColumn=False, 
                    showRowStripes=False,  # No row stripes for cleaner look
                    showColumnStripes=False  # No column stripes for cleaner look
                )
                table.tableStyleInfo = style
                
                # Add the table to the worksheet
                ws.add_table(table)
                if DEBUG:
                    print(f"Added Excel table formatting to '{sheet_name}' sheet")

                # Find URL column index
                header_row = ws[1]
                url_col_index = None
                for cell in header_row:
                    if cell.value == "URL":
                        url_col_index = cell.column
                        break
                
                if url_col_index:
                    # Check for special columns (Notes, NextGen, NextGen?) after URL
                    special_col_index = None
                    special_columns_to_check = ['Notes', 'NextGen', 'NextGen?']
                    
                    # Find the last special column after URL
                    for col_idx in range(url_col_index + 1, max_col + 1):
                        cell_value = ws.cell(row=1, column=col_idx).value
                        if cell_value in special_columns_to_check:
                            special_col_index = col_idx
                    
                    # Insert new column after the last special column if it exists, otherwise after URL
                    insert_position = special_col_index + 1 if special_col_index else url_col_index + 1
                    ws.insert_cols(insert_position)
                    
                    # Add header for link column
                    link_cell = ws.cell(row=1, column=insert_position)
                    link_cell.value = "Link"
                    
                    # Add hyperlink formula to each row
                    for row in range(2, max_row + 1):
                        formula_cell = ws.cell(row=row, column=insert_position)
                        url_cell = ws.cell(row=row, column=url_col_index).coordinate
                        formula_cell.value = f'=HYPERLINK({url_cell},"🔗")'
                    
                    # Hide URL column
                    col_letter = ws.cell(row=1, column=url_col_index).column_letter
                    ws.column_dimensions[col_letter].hidden = True
                    
                    # Adjust table range to include new column
                    table.ref = f"A1:{ws.cell(max_row, max_col + 1).coordinate}"
                    if DEBUG:
                        print(f"Added hyperlink column and hid URL column in '{sheet_name}' sheet")
    
    # Format Summary tab (Content Summary)
    if summary_tab_name in wb.sheetnames:
        ws = wb[summary_tab_name]
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Style first table headers
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Style second table headers (metadata)
        metadata_start_row = len(summary_df) + 4
        for cell in ws[metadata_start_row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Style third table headers (languages) if it exists
        if 'all_languages' in locals() and len(all_languages) > 0:
            lang_start_row = len(summary_df) + len(metadata_df) + 7
            for cell in ws[lang_start_row]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(excel_file_path)
    
    print(f"\nExcel analysis file created: {excel_file_path}")
    print(f"Tab 1: Complete Data ({len(df)} rows)")
    # if df_pivots is not None:
    #     print(f"Tab 3: Pivot Columns ({len(df_pivots)} rows with individual pivot boolean columns)")
    #     print(f"Tab 4: Content Summary with statistics tables")
    # else:
    print(f"Tab 2: Content Summary with statistics tables")
    
    return excel_file_path

if __name__ == "__main__":
    import sys
    
    # Check if user wants to create Excel analysis
    if len(sys.argv) > 1 and sys.argv[1] == "--excel":
        # If a CSV file path is provided as second argument, use it
        csv_path = sys.argv[2] if len(sys.argv) > 2 else None
        create_excel_analysis(csv_path)
    else:
        add_metadata_to_csv()
